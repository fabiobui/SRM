# Dashboard Views
import json
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Q, F
from django.core.serializers.json import DjangoJSONEncoder

from vendor_management_system.vendors.models import Vendor, Address, Category


@login_required
def vendor_dashboard_view(request):
    """
    Dashboard view che mostra statistiche e grafici sui fornitori
    """
    vendors = Vendor.objects.select_related('category', 'service_type', 'service_type__parent', 'address').prefetch_related('competences', 'vendor_documents__document_type').all()
    
    # Summary statistics
    total_vendors = vendors.count()
    active_vendors = vendors.filter(is_active=True).count()
    positive_vendors = vendors.filter(
        vendor_final_evaluation__in=['POSITIVO', 'MOLTO POSITIVO', 'Positivo', 'Molto Positivo']
    ).count()
    pending_qualification = vendors.filter(
        qualification_status__in=['PENDING', 'IN_PROGRESS', 'NOT_STARTED']
    ).count()
    high_risk_vendors = vendors.filter(risk_level__in=['HIGH', 'CRITICAL']).count()
    
    # Chart data aggregations
    chart_data = {
        'by_vendor_type': list(
            vendors.values('vendor_type')
            .annotate(count=Count('vendor_code'), type=F('vendor_type'))
            .order_by('-count')
        ),
        'by_ico_consultant': [
            {'is_ico': True, 'count': vendors.filter(is_ico_consultant=True).count()},
            {'is_ico': False, 'count': vendors.filter(is_ico_consultant=False).count()}
        ],
        'by_region': list(
            vendors.exclude(address__isnull=True)
            .exclude(address__region__isnull=True)
            .exclude(address__region='')
            .values('address__region')
            .annotate(count=Count('vendor_code'), region=F('address__region'))
            .order_by('-count')
        ),
        'by_service_type': list(
            vendors.values('service_type__name')
            .annotate(count=Count('vendor_code'), service_type=F('service_type__name'))
            .order_by('-count')
        ),
        'by_quality': [],
        'by_fulfillment': [],
        'by_competencies': [],
        'by_certifications': [],
    }
    
    # Add count for vendors without address or region
    vendors_no_region = vendors.filter(
        Q(address__isnull=True) | 
        Q(address__region__isnull=True) | 
        Q(address__region='')
    ).count()
    if vendors_no_region > 0:
        chart_data['by_region'].append({
            'region': 'Non specificato',
            'count': vendors_no_region
        })
    
    # Competencies aggregation
    try:
        from vendor_management_system.vendors.models import Competence
        
        competencies_data = (
            Competence.objects
            .annotate(vendor_count=Count('vendors_with_competence'))
            .filter(vendor_count__gt=0)
            .values('name', 'vendor_count')
            .order_by('-vendor_count')
        )
        
        for comp_data in competencies_data:
            chart_data['by_competencies'].append({
                'competency': comp_data['name'],
                'count': comp_data['vendor_count']
            })
        
    except Exception as e:
        pass
    
    # Se non ci sono competenze, mostra placeholder
    if not chart_data['by_competencies']:
        chart_data['by_competencies'] = [
            {'competency': 'Nessuna competenza assegnata', 'count': 0}
        ]
    
    # Certifications aggregation (filtra per has_certification=True)
    try:
        from vendor_management_system.vendors.models import Competence, VendorCompetence
        
        certifications_data = (
            Competence.objects
            .annotate(
                vendor_count=Count(
                    'vendor_assignments',
                    filter=Q(vendor_assignments__has_certification=True)
                )
            )
            .filter(vendor_count__gt=0)
            .values('name', 'vendor_count')
            .order_by('-vendor_count')
        )
        
        for cert_data in certifications_data:
            chart_data['by_certifications'].append({
                'certification': cert_data['name'],
                'count': cert_data['vendor_count']
            })
        
    except Exception as e:
        pass
    
    # Se non ci sono certificazioni, mostra placeholder
    if not chart_data['by_certifications']:
        chart_data['by_certifications'] = [
            {'certification': 'Nessuna certificazione assegnata', 'count': 0}
        ]
    
    # Quality rating distribution
    quality_ranges = [
        ('0-1', 0, 1),
        ('1-2', 1, 2),
        ('2-3', 2, 3),
        ('3-4', 3, 4),
        ('4-5', 4, 5),
    ]
    for range_label, min_val, max_val in quality_ranges:
        count = vendors.filter(
            quality_rating_avg__gte=min_val,
            quality_rating_avg__lt=max_val
        ).count()
        chart_data['by_quality'].append({'range': range_label, 'count': count})
    
    # Fulfillment rate distribution
    fulfillment_ranges = [
        ('0-20%', 0, 20),
        ('20-40%', 20, 40),
        ('40-60%', 40, 60),
        ('60-80%', 60, 80),
        ('80-100%', 80, 100),
    ]
    for range_label, min_val, max_val in fulfillment_ranges:
        count = vendors.filter(
            fulfillment_rate__gte=min_val,
            fulfillment_rate__lt=max_val
        ).count()
        chart_data['by_fulfillment'].append({'range': range_label, 'count': count})
    
    # Vendors data for table
    vendors_data = []
    for vendor in vendors:
        vendor_dict = {
            'vendor_code': vendor.vendor_code,
            'name': vendor.name,
            'email': vendor.email,
            'phone': vendor.phone if hasattr(vendor, 'phone') else None,
            'qualification_status': vendor.qualification_status,
            'risk_level': vendor.risk_level,
            'quality_rating_avg': vendor.quality_rating_avg,
            'fulfillment_rate': vendor.fulfillment_rate,
            'vat_number': vendor.vat_number,
            'fiscal_code': vendor.fiscal_code,
            'vendor_type': vendor.vendor_type,
            'is_ico_consultant': vendor.is_ico_consultant,
            'is_active': vendor.is_active,
            'vendor_final_evaluation': vendor.vendor_final_evaluation,
            'category': {'name': vendor.category.name if vendor.category else None},
            'service_type': {
                'name': vendor.service_type.name if vendor.service_type else None,
                'parent': vendor.service_type.parent.name if vendor.service_type and vendor.service_type.parent else None
            },
            'address': {
                'street_address': vendor.address.street_address if vendor.address else None,
                'city': vendor.address.city if vendor.address else None,
                'state_province': vendor.address.state_province if vendor.address else None,
                'region': vendor.address.region if vendor.address else None,
                'postal_code': vendor.address.postal_code if vendor.address else None,
                'country': vendor.address.country if vendor.address else 'Italia',
            } if vendor.address else None,
            'competences': [comp.name for comp in vendor.competences.all()],
            'certifications': [
                vc.competence.name 
                for vc in vendor.vendor_competences.filter(has_certification=True)
            ]
        }
        vendors_data.append(vendor_dict)
    
    context = {
        'total_vendors': total_vendors,
        'active_vendors': active_vendors,
        'positive_vendors': positive_vendors,
        'pending_qualification': pending_qualification,
        'high_risk_vendors': high_risk_vendors,
        'chart_data_json': json.dumps(chart_data, cls=DjangoJSONEncoder),
        'vendors_data_json': json.dumps(vendors_data, cls=DjangoJSONEncoder),
    }
    
    return render(request, 'vendors/vendor_dashboard.html', context)


@csrf_exempt
def dashboard_stats_api(request):
    """
    API endpoint per statistiche dashboard
    """
    vendors = Vendor.objects.all()
    
    stats = {
        'total': vendors.count(),
        'active': vendors.filter(is_active=True).count(),
        'pending_qualification': vendors.filter(
            qualification_status__in=['PENDING', 'IN_PROGRESS', 'NOT_STARTED']
        ).count(),
        'high_risk': vendors.filter(risk_level__in=['HIGH', 'CRITICAL']).count(),
    }
    
    return JsonResponse(stats)


@csrf_exempt
def dashboard_vendors_list_api(request):
    """
    API endpoint per lista fornitori con filtri
    """
    vendors = Vendor.objects.select_related('category', 'service_type').all()
    
    # Apply filters from query params
    category = request.GET.get('category')
    qualification_status = request.GET.get('qualification_status')
    risk_level = request.GET.get('risk_level')
    service_type = request.GET.get('service_type')
    search = request.GET.get('search', '')
    
    if category:
        vendors = vendors.filter(category__name=category)
    if qualification_status:
        vendors = vendors.filter(qualification_status=qualification_status)
    if risk_level:
        vendors = vendors.filter(risk_level=risk_level)
    if service_type:
        vendors = vendors.filter(service_type__name=service_type)
    if search:
        vendors = vendors.filter(
            Q(vendor_code__icontains=search) |
            Q(name__icontains=search) |
            Q(email__icontains=search) |
            Q(vat_number__icontains=search) |
            Q(fiscal_code__icontains=search)
        )
    
    vendors_data = list(vendors.values(
        'vendor_code', 'name', 'email',
        'qualification_status', 'risk_level',
        'quality_rating_avg', 'fulfillment_rate',
        'category__name', 'service_type__name'
    ))
    
    # Rename nested fields
    for vendor in vendors_data:
        vendor['category'] = {'name': vendor.pop('category__name', None)}
        vendor['service_type'] = {'name': vendor.pop('service_type__name', None)}
    
    return JsonResponse({'vendors': vendors_data})


@csrf_exempt
def export_vendors_excel(request):
    """
    Export vendors to Excel with applied filters
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from django.http import HttpResponse
    
    vendors = Vendor.objects.select_related('category', 'service_type').all()
    
    # Apply filters
    category = request.GET.get('category')
    qualification_status = request.GET.get('qualification_status')
    risk_level = request.GET.get('risk_level')
    service_type = request.GET.get('service_type')
    search = request.GET.get('search', '')
    
    if category:
        vendors = vendors.filter(category__name=category)
    if qualification_status:
        vendors = vendors.filter(qualification_status=qualification_status)
    if risk_level:
        vendors = vendors.filter(risk_level=risk_level)
    if service_type:
        vendors = vendors.filter(service_type__name=service_type)
    if search:
        vendors = vendors.filter(
            Q(vendor_code__icontains=search) |
            Q(name__icontains=search) |
            Q(email__icontains=search)
        )
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fornitori"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['Nome', 'Tipo', 'Email', 'Telefono', 'Regione', 'Provincia', 'Valutazione Complessiva']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Data rows
    for row, vendor in enumerate(vendors, 2):
        ws.cell(row=row, column=1, value=vendor.name)
        ws.cell(row=row, column=2, value=vendor.vendor_type)
        ws.cell(row=row, column=3, value=vendor.email)
        ws.cell(row=row, column=4, value=vendor.phone if hasattr(vendor, 'phone') else '')
        ws.cell(row=row, column=5, value=vendor.address.region if vendor.address else '')
        ws.cell(row=row, column=6, value=vendor.address.state_province if vendor.address else '')
        ws.cell(row=row, column=7, value=vendor.vendor_final_evaluation or 'DA VALUTARE')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=fornitori_export.xlsx'
    
    wb.save(response)
    return response


@login_required
def post_login_redirect(request):
    """
    Decide la destinazione post-login in base al profilo utente.
    - admin/staff -> Django admin
    - gruppo 'backoffice' -> vendor_dashboard
    - fallback -> vendor_dashboard
    """
    user = request.user
    if user.is_superuser or user.is_staff:
        return redirect('admin:index')
    if user.groups.filter(name__iexact='backoffice').exists():
        return redirect('vendor_dashboard')
    return redirect('vendor_dashboard')
