# Imports
import json
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Avg, Q, Case, When, F
from django.core.serializers.json import DjangoJSONEncoder
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from rest_framework import permissions, response, status, viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.authtoken.views import ObtainAuthToken

from vendor_management_system.core.authentication import (
    QueryParameterTokenAuthentication,
)

from vendor_management_system.core.serializers import QueryParamAuthTokenSerializer
from vendor_management_system.vendors.models import Vendor, Address, Category
from vendor_management_system.vendors.serializers import (
    VendorCreateUpdateSerializer,
    VendorSerializer,
    VendorListSerializer,
    VendorQualificationSerializer,
    VendorAuditSerializer,
    VendorPerformanceSerializer,
    AddressSerializer,
    AddressManagementSerializer,
    CategorySerializer,
    CategoryManagementSerializer,
    CategoryCompactSerializer,
    CategoryTreeSerializer,
    CategoryStatsSerializer,
)


# ViewSet per Address
class AddressViewSet(viewsets.ViewSet):
    """ViewSet for managing addresses independently"""
    permission_classes = [permissions.IsAuthenticated]
    authentication_classes = [QueryParameterTokenAuthentication]

    @swagger_auto_schema(
        operation_id="addresses--list",
        operation_description="List all addresses",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
            openapi.Parameter(
                name="country",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
                description="Filter by country",
            ),
            openapi.Parameter(
                name="city",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
                description="Filter by city",
            ),
            openapi.Parameter(
                name="address_type",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
                description="Filter by address type",
            ),
        ],
        responses={
            status.HTTP_200_OK: openapi.Response(
                "List of addresses", schema=AddressSerializer(many=True)
            ),
        },
        tags=["Addresses"],
    )
    def list(self, request):
        addresses = Address.objects.filter(is_active=True)
        
        # Apply filters
        country = request.query_params.get('country')
        if country:
            addresses = addresses.filter(country__icontains=country)
            
        city = request.query_params.get('city')
        if city:
            addresses = addresses.filter(city__icontains=city)
            
        address_type = request.query_params.get('address_type')
        if address_type:
            addresses = addresses.filter(address_type=address_type)

        serializer = AddressSerializer(addresses, many=True)
        return response.Response(serializer.data, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        operation_id="addresses--create",
        operation_description="Create a new address",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            )
        ],
        request_body=AddressManagementSerializer,
        responses={
            status.HTTP_201_CREATED: openapi.Response(
                "Created address", schema=AddressSerializer
            ),
        },
        tags=["Addresses"],
    )
    def create(self, request):
        serializer = AddressManagementSerializer(data=request.data)
        if serializer.is_valid():
            address = serializer.save()
            response_serializer = AddressSerializer(address)
            return response.Response(
                response_serializer.data, status=status.HTTP_201_CREATED
            )
        return response.Response(
            serializer.errors, status=status.HTTP_400_BAD_REQUEST
        )

    @swagger_auto_schema(
        operation_id="addresses--retrieve",
        operation_description="Retrieve a specific address",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        responses={
            status.HTTP_200_OK: openapi.Response(
                "Address details", schema=AddressSerializer
            ),
        },
        tags=["Addresses"],
    )
    def retrieve(self, request, address_id=None):
        address = get_object_or_404(Address, id=address_id)
        serializer = AddressSerializer(address)
        return response.Response(serializer.data, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        operation_id="addresses--update",
        operation_description="Update a specific address",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        request_body=AddressManagementSerializer,
        responses={
            status.HTTP_200_OK: openapi.Response(
                "Updated address", schema=AddressSerializer
            ),
        },
        tags=["Addresses"],
    )
    def update(self, request, address_id=None):
        address = get_object_or_404(Address, id=address_id)
        serializer = AddressManagementSerializer(
            address, data=request.data, partial=True
        )
        if serializer.is_valid():
            serializer.save()
            response_serializer = AddressSerializer(address)
            return response.Response(
                response_serializer.data, status=status.HTTP_200_OK
            )
        return response.Response(
            serializer.errors, status=status.HTTP_400_BAD_REQUEST
        )

    @swagger_auto_schema(
        operation_id="addresses--destroy",
        operation_description="Delete a specific address",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        responses={
            status.HTTP_204_NO_CONTENT: "Address deleted",
        },
        tags=["Addresses"],
    )
    def destroy(self, request, address_id=None):
        address = get_object_or_404(Address, id=address_id)
        address.delete()
        return response.Response(status=status.HTTP_204_NO_CONTENT)


# ViewSet per Category
class CategoryViewSet(viewsets.ViewSet):
    """ViewSet for managing categories"""
    permission_classes = [permissions.IsAuthenticated]
    authentication_classes = [QueryParameterTokenAuthentication]

    @swagger_auto_schema(
        operation_id="categories--list",
        operation_description="List all categories",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
            openapi.Parameter(
                name="is_active",
                format="boolean",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_BOOLEAN,
                required=False,
                description="Filter by active status",
            ),
            openapi.Parameter(
                name="parent",
                format="uuid",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
                description="Filter by parent category ID",
            ),
            openapi.Parameter(
                name="requires_certification",
                format="boolean",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_BOOLEAN,
                required=False,
                description="Filter by certification requirement",
            ),
        ],
        responses={
            status.HTTP_200_OK: openapi.Response(
                "List of categories", schema=CategorySerializer(many=True)
            ),
        },
        tags=["Categories"],
    )
    def list(self, request):
        categories = Category.objects.all().order_by('sort_order', 'name')
        
        # Apply filters
        is_active = request.query_params.get('is_active')
        if is_active is not None:
            categories = categories.filter(is_active=is_active.lower() == 'true')
            
        parent = request.query_params.get('parent')
        if parent:
            if parent.lower() == 'null':
                categories = categories.filter(parent__isnull=True)
            else:
                categories = categories.filter(parent__id=parent)
                
        requires_certification = request.query_params.get('requires_certification')
        if requires_certification is not None:
            categories = categories.filter(requires_certification=requires_certification.lower() == 'true')

        serializer = CategorySerializer(categories, many=True)
        return response.Response(serializer.data, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        operation_id="categories--create",
        operation_description="Create a new category",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            )
        ],
        request_body=CategoryManagementSerializer,
        responses={
            status.HTTP_201_CREATED: openapi.Response(
                "Created category", schema=CategorySerializer
            ),
        },
        tags=["Categories"],
    )
    def create(self, request):
        serializer = CategoryManagementSerializer(data=request.data)
        if serializer.is_valid():
            category = serializer.save()
            response_serializer = CategorySerializer(category)
            return response.Response(
                response_serializer.data, status=status.HTTP_201_CREATED
            )
        return response.Response(
            serializer.errors, status=status.HTTP_400_BAD_REQUEST
        )

    @swagger_auto_schema(
        operation_id="categories--retrieve",
        operation_description="Retrieve a specific category",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        responses={
            status.HTTP_200_OK: openapi.Response(
                "Category details", schema=CategorySerializer
            ),
        },
        tags=["Categories"],
    )
    def retrieve(self, request, category_id=None):
        category = get_object_or_404(Category, id=category_id)
        serializer = CategorySerializer(category)
        return response.Response(serializer.data, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        operation_id="categories--update",
        operation_description="Update a specific category",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        request_body=CategoryManagementSerializer,
        responses={
            status.HTTP_200_OK: openapi.Response(
                "Updated category", schema=CategorySerializer
            ),
        },
        tags=["Categories"],
    )
    def update(self, request, category_id=None):
        category = get_object_or_404(Category, id=category_id)
        serializer = CategoryManagementSerializer(
            category, data=request.data, partial=True
        )
        if serializer.is_valid():
            serializer.save()
            response_serializer = CategorySerializer(category)
            return response.Response(
                response_serializer.data, status=status.HTTP_200_OK
            )
        return response.Response(
            serializer.errors, status=status.HTTP_400_BAD_REQUEST
        )

    @swagger_auto_schema(
        operation_id="categories--destroy",
        operation_description="Delete a specific category",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        responses={
            status.HTTP_204_NO_CONTENT: "Category deleted",
            status.HTTP_400_BAD_REQUEST: "Cannot delete category with vendors",
        },
        tags=["Categories"],
    )
    def destroy(self, request, category_id=None):
        category = get_object_or_404(Category, id=category_id)
        
        # Check if category has vendors
        if category.vendors.exists():
            return response.Response(
                {"detail": "Cannot delete category that has vendors assigned. Please reassign vendors first."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if category has subcategories
        if category.subcategories.exists():
            return response.Response(
                {"detail": "Cannot delete category that has subcategories. Please delete subcategories first."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        category.delete()
        return response.Response(status=status.HTTP_204_NO_CONTENT)

    @swagger_auto_schema(
        operation_id="categories--tree",
        operation_description="Get category hierarchy tree",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        responses={
            status.HTTP_200_OK: openapi.Response(
                "Category tree", schema=CategoryTreeSerializer(many=True)
            ),
        },
        tags=["Categories"],
    )
    @action(detail=False, methods=['get'], url_path='tree')
    def tree(self, request):
        # Get only root categories (without parent)
        root_categories = Category.objects.filter(
            parent__isnull=True, 
            is_active=True
        ).order_by('sort_order', 'name')
        
        serializer = CategoryTreeSerializer(root_categories, many=True)
        return response.Response(serializer.data, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        operation_id="categories--stats",
        operation_description="Get category statistics",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        responses={
            status.HTTP_200_OK: openapi.Response(
                "Category statistics", schema=CategoryStatsSerializer(many=True)
            ),
        },
        tags=["Categories"],
    )
    @action(detail=False, methods=['get'], url_path='stats')
    def stats(self, request):
        categories = Category.objects.filter(is_active=True).order_by('sort_order', 'name')
        serializer = CategoryStatsSerializer(categories, many=True)
        return response.Response(serializer.data, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        operation_id="categories--vendors",
        operation_description="Get vendors in a specific category",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
            openapi.Parameter(
                name="include_subcategories",
                format="boolean",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_BOOLEAN,
                required=False,
                description="Include vendors from subcategories",
            ),
        ],
        responses={
            status.HTTP_200_OK: openapi.Response(
                "Vendors in category", schema=VendorListSerializer(many=True)
            ),
        },
        tags=["Categories"],
    )
    @action(detail=True, methods=['get'], url_path='vendors')
    def vendors(self, request, category_id=None):
        category = get_object_or_404(Category, id=category_id)
        include_subcategories = request.query_params.get('include_subcategories', 'false').lower() == 'true'
        
        if include_subcategories:
            # Get all descendant categories
            categories = [category] + category.get_descendants()
            vendors = Vendor.objects.filter(category__in=categories)
        else:
            vendors = category.vendors.all()
        
        vendors = vendors.order_by('name')
        serializer = VendorListSerializer(vendors, many=True)
        return response.Response(serializer.data, status=status.HTTP_200_OK)


# ViewSet per Vendor
class VendorViewSet(viewsets.ViewSet):
    """ViewSet for managing vendors"""
    permission_classes = [permissions.IsAuthenticated]
    authentication_classes = [QueryParameterTokenAuthentication]

    @swagger_auto_schema(
        operation_id="vendors--list-vendors",
        operation_description="List all vendors with basic information",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
            openapi.Parameter(
                name="qualification_status",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
                description="Filter by qualification status (PENDING, APPROVED, REJECTED)",
            ),
            openapi.Parameter(
                name="risk_level",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
                description="Filter by risk level (LOW, MEDIUM, HIGH)",
            ),
            openapi.Parameter(
                name="category",
                format="uuid",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
                description="Filter by category ID",
            ),
            openapi.Parameter(
                name="category_code",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
                description="Filter by category code",
            ),
            openapi.Parameter(
                name="requires_certification",
                format="boolean",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_BOOLEAN,
                required=False,
                description="Filter by certification requirement",
            ),
        ],
        responses={
            status.HTTP_200_OK: openapi.Response(
                "List of all vendors", schema=VendorListSerializer(many=True)
            ),
            status.HTTP_401_UNAUTHORIZED: "Unauthorized",
        },
        tags=["Vendors"],
    )
    def list(self, request):
        # Get all vendors con select_related per ottimizzare
        vendors = Vendor.objects.select_related('category', 'address').all()
        
        # Apply existing filters
        qualification_status = request.query_params.get('qualification_status')
        if qualification_status:
            vendors = vendors.filter(qualification_status=qualification_status)
            
        risk_level = request.query_params.get('risk_level')
        if risk_level:
            vendors = vendors.filter(risk_level=risk_level)
        
        # New category filters
        category = request.query_params.get('category')
        if category:
            vendors = vendors.filter(category__id=category)
            
        category_code = request.query_params.get('category_code')
        if category_code:
            vendors = vendors.filter(category__code__iexact=category_code)
            
        requires_certification = request.query_params.get('requires_certification')
        if requires_certification is not None:
            vendors = vendors.filter(category__requires_certification=requires_certification.lower() == 'true')

        # Serialize the vendors using lightweight serializer
        serializer = VendorListSerializer(vendors, many=True)

        # Return the response
        return response.Response(serializer.data, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        operation_id="vendors--create-vendor",
        operation_description="Create a new vendor",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            )
        ],
        request_body=VendorCreateUpdateSerializer,
        responses={
            status.HTTP_201_CREATED: openapi.Response(
                "The created vendor", schema=VendorSerializer
            ),
            status.HTTP_400_BAD_REQUEST: "Bad request",
            status.HTTP_401_UNAUTHORIZED: "Unauthorized",
        },
        tags=["Vendors"],
    )
    def create(self, request):
        # Deserialize and validate the request data
        vendor_create_serializer = VendorCreateUpdateSerializer(data=request.data)

        # If the provided data is valid
        if vendor_create_serializer.is_valid():
            # Create a new Vendor object
            vendor = vendor_create_serializer.save()

            # Serialize the vendor with full details
            vendor_serializer = VendorSerializer(vendor)

            # Return the response
            return response.Response(
                vendor_serializer.data, status=status.HTTP_201_CREATED
            )

        # Return the error response
        return response.Response(
            vendor_create_serializer.errors, status=status.HTTP_400_BAD_REQUEST
        )

    @swagger_auto_schema(
        operation_id="vendors--retrieve-vendor",
        operation_description="Retrieve a specific vendor's complete details",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
            openapi.Parameter(
                name="vendor_code",
                format="string",
                in_=openapi.IN_PATH,
                type=openapi.TYPE_STRING,
                required=True,
                description="The vendor_code for the vendor to retrieve",
            ),
        ],
        responses={
            status.HTTP_200_OK: openapi.Response(
                "The retrieved vendor", schema=VendorSerializer
            ),
            status.HTTP_404_NOT_FOUND: "Vendor not found",
            status.HTTP_401_UNAUTHORIZED: "Unauthorized",
        },
        tags=["Vendors"],
    )
    def retrieve(self, request, vendor_code=None):
        # Get the vendor by vendor_code
        vendor = get_object_or_404(Vendor, vendor_code=vendor_code)

        # Serialize the vendor with full details
        serializer = VendorSerializer(vendor)

        # Return the response
        return response.Response(serializer.data, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        operation_id="vendors--update-vendor",
        operation_description="Update a specific vendor's basic information",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
            openapi.Parameter(
                name="vendor_code",
                format="string",
                in_=openapi.IN_PATH,
                type=openapi.TYPE_STRING,
                required=True,
                description="The vendor_code for the vendor to update",
            ),
        ],
        request_body=VendorCreateUpdateSerializer,
        responses={
            status.HTTP_200_OK: openapi.Response(
                "The updated vendor", schema=VendorSerializer
            ),
            status.HTTP_400_BAD_REQUEST: "Bad request",
            status.HTTP_404_NOT_FOUND: "Vendor not found",
            status.HTTP_401_UNAUTHORIZED: "Unauthorized",
        },
        tags=["Vendors"],
    )
    def update(self, request, vendor_code=None):
        # Get the vendor by vendor_code
        vendor = get_object_or_404(Vendor, vendor_code=vendor_code)

        # Deserialize and validate the data
        vendor_update_serializer = VendorCreateUpdateSerializer(
            vendor, data=request.data, partial=True
        )

        # If the provided data is valid
        if vendor_update_serializer.is_valid():
            # Save the vendor
            vendor_update_serializer.save()

            # Fetch the updated vendor
            vendor = Vendor.objects.get(vendor_code=vendor_code)

            # Serialize the vendor
            serializer = VendorSerializer(vendor)

            # Return the response
            return response.Response(serializer.data, status=status.HTTP_200_OK)

        # Return the error response
        return response.Response(
            vendor_update_serializer.errors, status=status.HTTP_400_BAD_REQUEST
        )

    @swagger_auto_schema(
        operation_id="vendors--destroy-vendor",
        operation_description="Delete a specific vendor",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
            openapi.Parameter(
                name="vendor_code",
                format="string",
                in_=openapi.IN_PATH,
                type=openapi.TYPE_STRING,
                required=True,
                description="The vendor_code for the vendor to delete",
            ),
        ],
        responses={
            status.HTTP_204_NO_CONTENT: "Vendor deleted",
            status.HTTP_404_NOT_FOUND: "Vendor not found",
            status.HTTP_401_UNAUTHORIZED: "Unauthorized",
        },
        tags=["Vendors"],
    )
    def destroy(self, request, vendor_code=None):
        # Get the vendor by vendor_code
        vendor = get_object_or_404(Vendor, vendor_code=vendor_code)

        # Delete the vendor
        vendor.delete()

        # Return the response
        return response.Response(status=status.HTTP_204_NO_CONTENT)

    # Custom action for qualification management
    @swagger_auto_schema(
        methods=['get'],
        operation_id="vendors--get-qualification",
        operation_description="Get vendor qualification details",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        responses={
            status.HTTP_200_OK: openapi.Response(
                "Vendor qualification details", schema=VendorQualificationSerializer
            ),
        },
        tags=["Vendor Qualification"],
    )
    @swagger_auto_schema(
        methods=['patch'],
        operation_id="vendors--update-qualification",
        operation_description="Update vendor qualification",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        request_body=VendorQualificationSerializer,
        responses={
            status.HTTP_200_OK: openapi.Response(
                "Updated qualification", schema=VendorQualificationSerializer
            ),
        },
        tags=["Vendor Qualification"],
    )
    @action(detail=True, methods=['get', 'patch'], url_path='qualification')
    def qualification(self, request, vendor_code=None):
        vendor = get_object_or_404(Vendor, vendor_code=vendor_code)
        
        if request.method == 'GET':
            serializer = VendorQualificationSerializer(vendor)
            return response.Response(serializer.data, status=status.HTTP_200_OK)
        
        elif request.method == 'PATCH':
            serializer = VendorQualificationSerializer(
                vendor, data=request.data, partial=True
            )
            if serializer.is_valid():
                serializer.save()
                return response.Response(serializer.data, status=status.HTTP_200_OK)
            return response.Response(
                serializer.errors, status=status.HTTP_400_BAD_REQUEST
            )

    # Custom action for audit management
    @swagger_auto_schema(
        methods=['get'],
        operation_id="vendors--get-audit",
        operation_description="Get vendor audit details",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        responses={
            status.HTTP_200_OK: openapi.Response(
                "Vendor audit details", schema=VendorAuditSerializer
            ),
        },
        tags=["Vendor Audit"],
    )
    @swagger_auto_schema(
        methods=['patch'],
        operation_id="vendors--update-audit",
        operation_description="Update vendor audit information",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        request_body=VendorAuditSerializer,
        responses={
            status.HTTP_200_OK: openapi.Response(
                "Updated audit info", schema=VendorAuditSerializer
            ),
        },
        tags=["Vendor Audit"],
    )
    @action(detail=True, methods=['get', 'patch'], url_path='audit')
    def audit(self, request, vendor_code=None):
        vendor = get_object_or_404(Vendor, vendor_code=vendor_code)
        
        if request.method == 'GET':
            serializer = VendorAuditSerializer(vendor)
            return response.Response(serializer.data, status=status.HTTP_200_OK)
        
        elif request.method == 'PATCH':
            serializer = VendorAuditSerializer(
                vendor, data=request.data, partial=True
            )
            if serializer.is_valid():
                serializer.save()
                return response.Response(serializer.data, status=status.HTTP_200_OK)
            return response.Response(
                serializer.errors, status=status.HTTP_400_BAD_REQUEST
            )

    # Custom action for performance management
    @swagger_auto_schema(
        methods=['get'],
        operation_id="vendors--get-performance",
        operation_description="Get vendor performance metrics",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        responses={
            status.HTTP_200_OK: openapi.Response(
                "Vendor performance metrics", schema=VendorPerformanceSerializer
            ),
        },
        tags=["Vendor Performance"],
    )
    @swagger_auto_schema(
        methods=['patch'],
        operation_id="vendors--update-performance",
        operation_description="Update vendor performance metrics",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        request_body=VendorPerformanceSerializer,
        responses={
            status.HTTP_200_OK: openapi.Response(
                "Updated performance metrics", schema=VendorPerformanceSerializer
            ),
        },
        tags=["Vendor Performance"],
    )
    @action(detail=True, methods=['get', 'patch'], url_path='performance')
    def performance(self, request, vendor_code=None):
        vendor = get_object_or_404(Vendor, vendor_code=vendor_code)
        
        if request.method == 'GET':
            serializer = VendorPerformanceSerializer(vendor)
            return response.Response(serializer.data, status=status.HTTP_200_OK)
        
        elif request.method == 'PATCH':
            serializer = VendorPerformanceSerializer(
                vendor, data=request.data, partial=True
            )
            if serializer.is_valid():
                serializer.save()
                return response.Response(serializer.data, status=status.HTTP_200_OK)
            return response.Response(
                serializer.errors, status=status.HTTP_400_BAD_REQUEST
            )

    # Custom action to get vendors requiring attention
    @swagger_auto_schema(
        operation_id="vendors--get-alerts",
        operation_description="Get vendors requiring attention (overdue audits, expired qualifications, etc.)",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        responses={
            status.HTTP_200_OK: openapi.Response(
                "Vendors requiring attention", 
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "overdue_audits": openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(
                                type=openapi.TYPE_OBJECT,
                                properties={
                                    "vendor_code": openapi.Schema(type=openapi.TYPE_STRING),
                                    "name": openapi.Schema(type=openapi.TYPE_STRING),
                                    "email": openapi.Schema(type=openapi.TYPE_STRING),
                                    "phone": openapi.Schema(type=openapi.TYPE_STRING),
                                    "category": openapi.Schema(type=openapi.TYPE_STRING),
                                    "qualification_status": openapi.Schema(type=openapi.TYPE_STRING),
                                    "risk_level": openapi.Schema(type=openapi.TYPE_STRING),
                                    "is_qualified": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                    "audit_overdue": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                }
                            )
                        ),
                        "expired_qualifications": openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(
                                type=openapi.TYPE_OBJECT,
                                properties={
                                    "vendor_code": openapi.Schema(type=openapi.TYPE_STRING),
                                    "name": openapi.Schema(type=openapi.TYPE_STRING),
                                    "email": openapi.Schema(type=openapi.TYPE_STRING),
                                    "phone": openapi.Schema(type=openapi.TYPE_STRING),
                                    "category": openapi.Schema(type=openapi.TYPE_STRING),
                                    "qualification_status": openapi.Schema(type=openapi.TYPE_STRING),
                                    "risk_level": openapi.Schema(type=openapi.TYPE_STRING),
                                    "is_qualified": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                    "audit_overdue": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                }
                            )
                        ),
                        "high_risk_vendors": openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(
                                type=openapi.TYPE_OBJECT,
                                properties={
                                    "vendor_code": openapi.Schema(type=openapi.TYPE_STRING),
                                    "name": openapi.Schema(type=openapi.TYPE_STRING),
                                    "email": openapi.Schema(type=openapi.TYPE_STRING),
                                    "phone": openapi.Schema(type=openapi.TYPE_STRING),
                                    "category": openapi.Schema(type=openapi.TYPE_STRING),
                                    "qualification_status": openapi.Schema(type=openapi.TYPE_STRING),
                                    "risk_level": openapi.Schema(type=openapi.TYPE_STRING),
                                    "is_qualified": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                    "audit_overdue": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                }
                            )
                        ),
                        "missing_certification": openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(
                                type=openapi.TYPE_OBJECT,
                                properties={
                                    "vendor_code": openapi.Schema(type=openapi.TYPE_STRING),
                                    "name": openapi.Schema(type=openapi.TYPE_STRING),
                                    "email": openapi.Schema(type=openapi.TYPE_STRING),
                                    "phone": openapi.Schema(type=openapi.TYPE_STRING),
                                    "category": openapi.Schema(type=openapi.TYPE_STRING),
                                    "qualification_status": openapi.Schema(type=openapi.TYPE_STRING),
                                    "risk_level": openapi.Schema(type=openapi.TYPE_STRING),
                                    "is_qualified": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                    "audit_overdue": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                }
                            )
                        ),
                        "no_category": openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(
                                type=openapi.TYPE_OBJECT,
                                properties={
                                    "vendor_code": openapi.Schema(type=openapi.TYPE_STRING),
                                    "name": openapi.Schema(type=openapi.TYPE_STRING),
                                    "email": openapi.Schema(type=openapi.TYPE_STRING),
                                    "phone": openapi.Schema(type=openapi.TYPE_STRING),
                                    "category": openapi.Schema(type=openapi.TYPE_STRING),
                                    "qualification_status": openapi.Schema(type=openapi.TYPE_STRING),
                                    "risk_level": openapi.Schema(type=openapi.TYPE_STRING),
                                    "is_qualified": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                    "audit_overdue": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                }
                            )
                        ),
                    }
                )
            ),
        },
        tags=["Vendor Alerts"],
    )
    @action(detail=False, methods=['get'], url_path='alerts')
    def alerts(self, request):
        from django.utils import timezone
        today = timezone.now().date()
        
        # Get vendors with overdue audits
        overdue_audits = Vendor.objects.select_related('category', 'address').filter(
            next_audit_due__lt=today
        ).exclude(next_audit_due__isnull=True)
        
        # Get vendors with expired qualifications
        expired_qualifications = Vendor.objects.select_related('category', 'address').filter(
            qualification_expiry__lt=today,
            qualification_status='APPROVED'
        ).exclude(qualification_expiry__isnull=True)
        
        # Get high risk vendors
        high_risk_vendors = Vendor.objects.select_related('category', 'address').filter(risk_level='HIGH')
        
        # Get vendors in categories requiring certification but without proper status
        missing_certification = Vendor.objects.select_related('category', 'address').filter(
            category__requires_certification=True,
            qualification_status__in=['PENDING', 'REJECTED']
        )
        
        # Get vendors without category assigned
        no_category = Vendor.objects.select_related('address').filter(category__isnull=True)
        
        data = {
            'overdue_audits': VendorListSerializer(overdue_audits, many=True).data,
            'expired_qualifications': VendorListSerializer(expired_qualifications, many=True).data,
            'high_risk_vendors': VendorListSerializer(high_risk_vendors, many=True).data,
            'missing_certification': VendorListSerializer(missing_certification, many=True).data,
            'no_category': VendorListSerializer(no_category, many=True).data,
        }
        
        return response.Response(data, status=status.HTTP_200_OK)

    # Address management for vendors
    @swagger_auto_schema(
        methods=['get'],
        operation_id="vendors--get-address",
        operation_description="Get vendor address",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        responses={
            status.HTTP_200_OK: openapi.Response(
                "Vendor address", schema=AddressSerializer
            ),
            status.HTTP_404_NOT_FOUND: "Address not found",
        },
        tags=["Vendor Address"],
    )
    @swagger_auto_schema(
        methods=['post'],
        operation_id="vendors--create-address",
        operation_description="Create address for vendor",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        request_body=AddressManagementSerializer,
        responses={
            status.HTTP_201_CREATED: openapi.Response(
                "Created address", schema=AddressSerializer
            ),
        },
        tags=["Vendor Address"],
    )
    @swagger_auto_schema(
        methods=['put'],
        operation_id="vendors--update-address",
        operation_description="Update vendor address",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        request_body=AddressManagementSerializer,
        responses={
            status.HTTP_200_OK: openapi.Response(
                "Updated address", schema=AddressSerializer
            ),
        },
        tags=["Vendor Address"],
    )
    @swagger_auto_schema(
        methods=['delete'],
        operation_id="vendors--delete-address",
        operation_description="Delete vendor address",
        manual_parameters=[
            openapi.Parameter(
                name="token",
                format="string",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="The token to authenticate the user",
            ),
        ],
        responses={
            status.HTTP_204_NO_CONTENT: "Address deleted",
        },
        tags=["Vendor Address"],
    )
    @action(detail=True, methods=['get', 'post', 'put', 'delete'], url_path='address')
    def address_management(self, request, vendor_code=None):
        vendor = get_object_or_404(Vendor, vendor_code=vendor_code)
        
        if request.method == 'GET':
            if vendor.address:
                serializer = AddressSerializer(vendor.address)
                return response.Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return response.Response(
                    {"detail": "Vendor has no address"}, 
                    status=status.HTTP_404_NOT_FOUND
                )
        
        elif request.method == 'POST':
            if vendor.address:
                return response.Response(
                    {"detail": "Vendor already has an address. Use PUT to update."}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            serializer = AddressManagementSerializer(data=request.data)
            if serializer.is_valid():
                address = serializer.save()
                vendor.address = address
                vendor.save()
                response_serializer = AddressSerializer(address)
                return response.Response(
                    response_serializer.data, status=status.HTTP_201_CREATED
                )
            return response.Response(
                serializer.errors, status=status.HTTP_400_BAD_REQUEST
            )
        
        elif request.method == 'PUT':
            if not vendor.address:
                return response.Response(
                    {"detail": "Vendor has no address. Use POST to create."}, 
                    status=status.HTTP_404_NOT_FOUND
                )
            
            serializer = AddressManagementSerializer(
                vendor.address, data=request.data, partial=True
            )
            if serializer.is_valid():
                serializer.save()
                response_serializer = AddressSerializer(vendor.address)
                return response.Response(
                    response_serializer.data, status=status.HTTP_200_OK
                )
            return response.Response(
                serializer.errors, status=status.HTTP_400_BAD_REQUEST
            )
        
        elif request.method == 'DELETE':
            if vendor.address:
                vendor.address.delete()
                vendor.address = None
                vendor.save()
                return response.Response(status=status.HTTP_204_NO_CONTENT)
            else:
                return response.Response(
                    {"detail": "Vendor has no address"}, 
                    status=status.HTTP_404_NOT_FOUND
                )


# Custom view to obtain an auth token
class QueryParamObtainAuthToken(ObtainAuthToken):
    serializer_class = QueryParamAuthTokenSerializer

    @swagger_auto_schema(
        operation_id="core--obtain-auth-token",
        operation_description="Obtain an auth token for a user",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["email", "password"],
            properties={
                "email": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description="Email address of the user",
                    format="email",
                ),
                "password": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description="Password of the user",
                    format="password",
                ),
            },
        ),
        responses={
            status.HTTP_200_OK: openapi.Response(
                "The auth token",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={"token": openapi.Schema(type=openapi.TYPE_STRING)},
                ),
            ),
            status.HTTP_400_BAD_REQUEST: "Bad request",
        },
        tags=["Rest API Authentication"],
    )
    def post(self, request, *args, **kwargs):
        return super().post(request, *args, **kwargs)


@login_required
def dashboard_redirect(request):
    """Reindirizza l'utente alla dashboard appropriata in base al ruolo"""
    user = request.user
    
    if user.is_admin():
        return redirect('admin-dashboard')
    elif user.is_bo_user():
        return redirect('backoffice-dashboard')
    elif user.is_vendor_user():
        return redirect('vendor-portal')
    else:
        messages.warning(request, "Il tuo account non ha un ruolo assegnato. Contatta l'amministratore.")
        return HttpResponse("""
        <div style="padding: 20px; font-family: Arial;">
            <h2>⚠️ Ruolo non assegnato</h2>
            <p>Il tuo account non ha un ruolo configurato.</p>
            <p>Contatta l'amministratore del sistema.</p>
            <a href="/admin/logout/">Logout</a>
        </div>
        """)


@login_required
def vendor_dashboard_view(request):
    """
    Dashboard view che mostra statistiche e grafici sui fornitori
    """
    vendors = Vendor.objects.select_related('category', 'service_type', 'service_type__parent', 'address').prefetch_related('competences').all()
    
    # Summary statistics
    total_vendors = vendors.count()
    active_vendors = vendors.filter(is_active=True).count()
    pending_qualification = vendors.filter(
        qualification_status__in=['PENDING', 'IN_PROGRESS', 'NOT_STARTED']
    ).count()
    high_risk_vendors = vendors.filter(risk_level__in=['HIGH', 'CRITICAL']).count()
    
    # Chart data aggregations
    chart_data = {
        'by_vendor_type': list(
            vendors.values('vendor_type')
            .annotate(count=Count('vendor_code'), type=F('vendor_type'))
            .order_by('-count')
        ),
        'by_service_type_parent': list(
            vendors.exclude(service_type__isnull=True)
            .exclude(service_type__parent__isnull=True)
            .values('service_type__parent__name')
            .annotate(count=Count('vendor_code'), service_type_parent=F('service_type__parent__name'))
            .order_by('-count')
        ),
        'by_region': list(
            vendors.exclude(address__isnull=True)
            .exclude(address__region__isnull=True)
            .exclude(address__region='')
            .values('address__region')
            .annotate(count=Count('vendor_code'), region=F('address__region'))
            .order_by('-count')
        ),
        'by_service_type': list(
            vendors.values('service_type__name')
            .annotate(count=Count('vendor_code'), service_type=F('service_type__name'))
            .order_by('-count')
        ),
        'by_quality': [],
        'by_fulfillment': [],
        'by_competencies': [],
    }
    
    # Add count for vendors without service_type parent
    vendors_no_parent = vendors.filter(
        Q(service_type__isnull=True) | Q(service_type__parent__isnull=True)
    ).count()
    if vendors_no_parent > 0:
        chart_data['by_service_type_parent'].append({
            'service_type_parent': 'Non specificato',
            'count': vendors_no_parent
        })
    
    # Add count for vendors without address or region
    vendors_no_region = vendors.filter(
        Q(address__isnull=True) | 
        Q(address__region__isnull=True) | 
        Q(address__region='')
    ).count()
    if vendors_no_region > 0:
        chart_data['by_region'].append({
            'region': 'Non specificato',
            'count': vendors_no_region
        })
    
    # Competencies aggregation
    try:
        from vendor_management_system.vendors.models import Competence
        
        competencies_data = (
            Competence.objects
            .annotate(vendor_count=Count('vendors_with_competence'))
            .filter(vendor_count__gt=0)
            .values('name', 'vendor_count')
            .order_by('-vendor_count')[:10]
        )
        
        for comp_data in competencies_data:
            chart_data['by_competencies'].append({
                'competency': comp_data['name'],
                'count': comp_data['vendor_count']
            })
        
    except Exception as e:
        pass
    
    # Se non ci sono competenze, mostra placeholder
    if not chart_data['by_competencies']:
        chart_data['by_competencies'] = [
            {'competency': 'Nessuna competenza assegnata', 'count': 0}
        ]
    
    # Quality rating distribution
    quality_ranges = [
        ('0-1', 0, 1),
        ('1-2', 1, 2),
        ('2-3', 2, 3),
        ('3-4', 3, 4),
        ('4-5', 4, 5),
    ]
    for range_label, min_val, max_val in quality_ranges:
        count = vendors.filter(
            quality_rating_avg__gte=min_val,
            quality_rating_avg__lt=max_val
        ).count()
        chart_data['by_quality'].append({'range': range_label, 'count': count})
    
    # Fulfillment rate distribution
    fulfillment_ranges = [
        ('0-20%', 0, 20),
        ('20-40%', 20, 40),
        ('40-60%', 40, 60),
        ('60-80%', 60, 80),
        ('80-100%', 80, 100),
    ]
    for range_label, min_val, max_val in fulfillment_ranges:
        count = vendors.filter(
            fulfillment_rate__gte=min_val,
            fulfillment_rate__lt=max_val
        ).count()
        chart_data['by_fulfillment'].append({'range': range_label, 'count': count})
    
    # Vendors data for table
    vendors_data = []
    for vendor in vendors:
        vendor_dict = {
            'vendor_code': vendor.vendor_code,
            'name': vendor.name,
            'email': vendor.email,
            'qualification_status': vendor.qualification_status,
            'risk_level': vendor.risk_level,
            'quality_rating_avg': vendor.quality_rating_avg,
            'fulfillment_rate': vendor.fulfillment_rate,
            'vat_number': vendor.vat_number,
            'fiscal_code': vendor.fiscal_code,
            'vendor_type': vendor.vendor_type,
            'category': {'name': vendor.category.name if vendor.category else None},
            'service_type': {
                'name': vendor.service_type.name if vendor.service_type else None,
                'parent': vendor.service_type.parent.name if vendor.service_type and vendor.service_type.parent else None
            },
            'address': {
                'street_address': vendor.address.street_address if vendor.address else None,
                'city': vendor.address.city if vendor.address else None,
                'state_province': vendor.address.state_province if vendor.address else None,
                'region': vendor.address.region if vendor.address else None,
                'postal_code': vendor.address.postal_code if vendor.address else None,
                'country': vendor.address.country if vendor.address else 'Italia',
            } if vendor.address else None,
            'competences': [comp.name for comp in vendor.competences.all()]
        }
        vendors_data.append(vendor_dict)
    
    context = {
        'total_vendors': total_vendors,
        'active_vendors': active_vendors,
        'pending_qualification': pending_qualification,
        'high_risk_vendors': high_risk_vendors,
        'chart_data_json': json.dumps(chart_data, cls=DjangoJSONEncoder),
        'vendors_data_json': json.dumps(vendors_data, cls=DjangoJSONEncoder),
    }
    
    return render(request, 'vendors/vendor_dashboard.html', context)


from django.http import JsonResponse
from django.db.models import Count, Avg, Q, Case, When, IntegerField


@csrf_exempt
def dashboard_stats_api(request):
    """
    API endpoint per statistiche dashboard
    """
    vendors = Vendor.objects.all()
    
    stats = {
        'total': vendors.count(),
        'active': vendors.filter(is_active=True).count(),
        'pending_qualification': vendors.filter(
            qualification_status__in=['PENDING', 'IN_PROGRESS', 'NOT_STARTED']
        ).count(),
        'high_risk': vendors.filter(risk_level__in=['HIGH', 'CRITICAL']).count(),
    }
    
    return JsonResponse(stats)


@csrf_exempt
def dashboard_vendors_list_api(request):
    """
    API endpoint per lista fornitori con filtri
    """
    vendors = Vendor.objects.select_related('category', 'service_type').all()
    
    # Apply filters from query params
    category = request.GET.get('category')
    qualification_status = request.GET.get('qualification_status')
    risk_level = request.GET.get('risk_level')
    service_type = request.GET.get('service_type')
    search = request.GET.get('search', '')
    
    if category:
        vendors = vendors.filter(category__name=category)
    if qualification_status:
        vendors = vendors.filter(qualification_status=qualification_status)
    if risk_level:
        vendors = vendors.filter(risk_level=risk_level)
    if service_type:
        vendors = vendors.filter(service_type__name=service_type)
    if search:
        vendors = vendors.filter(
            Q(vendor_code__icontains=search) |
            Q(name__icontains=search) |
            Q(email__icontains=search) |
            Q(vat_number__icontains=search) |
            Q(fiscal_code__icontains=search)
        )
    
    vendors_data = list(vendors.values(
        'vendor_code', 'name', 'email',
        'qualification_status', 'risk_level',
        'quality_rating_avg', 'fulfillment_rate',
        'category__name', 'service_type__name'
    ))
    
    # Rename nested fields
    for vendor in vendors_data:
        vendor['category'] = {'name': vendor.pop('category__name', None)}
        vendor['service_type'] = {'name': vendor.pop('service_type__name', None)}
    
    return JsonResponse({'vendors': vendors_data})


@csrf_exempt
def export_vendors_excel(request):
    """
    Export vendors to Excel with applied filters
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from django.http import HttpResponse
    
    vendors = Vendor.objects.select_related('category', 'service_type').all()
    
    # Apply filters
    category = request.GET.get('category')
    qualification_status = request.GET.get('qualification_status')
    risk_level = request.GET.get('risk_level')
    service_type = request.GET.get('service_type')
    search = request.GET.get('search', '')
    
    if category:
        vendors = vendors.filter(category__name=category)
    if qualification_status:
        vendors = vendors.filter(qualification_status=qualification_status)
    if risk_level:
        vendors = vendors.filter(risk_level=risk_level)
    if service_type:
        vendors = vendors.filter(service_type__name=service_type)
    if search:
        vendors = vendors.filter(
            Q(vendor_code__icontains=search) |
            Q(name__icontains=search) |
            Q(email__icontains=search)
        )
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fornitori"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['Codice', 'Nome', 'Email', 'Categoria', 'Stato Qualifica', 
               'Rischio', 'Tipo Servizio', 'Rating Qualità', 'Tasso Adempimento']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Data rows
    for row, vendor in enumerate(vendors, 2):
        ws.cell(row=row, column=1, value=vendor.vendor_code)
        ws.cell(row=row, column=2, value=vendor.name)
        ws.cell(row=row, column=3, value=vendor.email)
        ws.cell(row=row, column=4, value=vendor.category.name if vendor.category else '')
        ws.cell(row=row, column=5, value=vendor.qualification_status)
        ws.cell(row=row, column=6, value=vendor.risk_level)
        ws.cell(row=row, column=7, value=vendor.service_type.name if vendor.service_type else '')
        ws.cell(row=row, column=8, value=vendor.quality_rating_avg or 0)
        ws.cell(row=row, column=9, value=vendor.fulfillment_rate or 0)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=fornitori_export.xlsx'
    
    wb.save(response)
    return response